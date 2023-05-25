# Convertiert  ein Toolbox Projekt zu einer CSV/ Excel Fähigen Datei format um

import sys, getopt, os
import regex as re
import csv, json
import pandas, openpyxl
import itertools
from bs4 import BeautifulSoup
from copy import copy, deepcopy
from openpyxl.utils.dataframe import dataframe_to_rows
from globals import *
from fileParser import getFilterFrom
from utils import *

filters = None

log = []

# gibt bei do_check und geladenen Wörterbüchern das korrigierte Wort zurück. Wenn die Annotationen eindeutig sind, werden sie automatisch aufgefüllt, wenn nicht, bleiben sie unangetastet. Für den Fall, dass Annotationen vollkommen fehlen, können diese automatisch aufgefüllt werden, deswegen gibt die Funktion immer eine Liste von Werten zurück, die mit extend() angefügt wird.
spannenindex = {}

# ...
raw_xml = {}

markers = {}
db_words = {}


def decode_toolbox_map(string, marker):
    def get_line(string, marker):
        string = re.sub("\n(?!\\\\)", " ", string)
        search_reg = "\\\\({}) ?(.+?(\n|$))?"
        this_line = re.search(search_reg.format(marker), string)
        return {marker: this_line.group(2) if this_line else ""}

    def next_line(string, marker):
        marker_stack.append(marker)

        if "mkrFollowingThis" in markers[marker]:
            new_marker = markers[marker]["mkrFollowingThis"]

            # in manchen Dateien ist als following auf den Übersetzungsmarker wieder ref eingegeben, um die Datei leicht zu erweitern. Das führt zu einem unendlichen Loop, der hier unterbrochen wird.
            if new_marker in marker_stack:
                marker_stack.pop()
                return get_line(string, marker)

        else:  # der letzte Marker ist die Übersetzung
            marker_stack.pop()
            return get_line(string, marker)

        nnn = next_line(string, new_marker)
        ttt = get_line(string, marker)

        ttt.update(nnn)

        marker_stack.pop()
        return ttt

    def next_block(string, marker):
        marker_stack.append(marker)

        if marker in string:
            this_line = re.search("\\\\({}) ?(.+?)?\n".format(marker), string)
            this_annotation = this_line.group(2)
        else:
            marker_stack.pop()
            return

        if "mkrFollowingThis" in markers[marker]:
            new_marker = markers[marker]["mkrFollowingThis"]

            if new_marker in marker_stack:
                input(marker_stack)
        else:
            # print(marker, this_annotation, "$")
            marker_stack.pop()
            return

        if not "jumps" in markers[
            marker]:  # ich verwende das vorhandensein von Lookup-Funktionen als Indikator dafür, dass wir es nicht mit einem Id- oder Record-Marker zu tun haben, sondern mit einem zu interlinearisierenden Eintrag. Das setzt voraus, dass auch in den Datenbanken solche Lookup-Funktionen beim höchsten Eintrag angesetzt werden, ansonsten werden die Wörterbucheinträge nicht interlinearisiert dargestellt
            aaa = {marker: {}}

            new_subset = re.split("(?V1)(?=\\\\{})".format(marker), string)
            for sub_string in new_subset[1:]:
                if marker in sub_string:
                    this_line = re.search("\\\\({}) ?(.+?)?\n".format(marker), sub_string)

                    this_annotation = this_line.group(2)
                else:
                    marker_stack.pop()
                    return

                map = next_block(sub_string, new_marker)

                aaa[marker][this_annotation] = map

            marker_stack.pop()
            return aaa
        else:
            aaa = []

            new_subset = re.split("(?V1)(?=\\\\{})".format(marker), string)
            for sub_string in new_subset[1:]:
                aaa.append(next_line(sub_string, marker))

            marker_stack.pop()
            return aaa

    return next_block(string, marker)


def decode_toolbox_json(map, marker, prefix):  # prefix = { marker : marker_value }	 (fName, id, rec)
    global words, log

    def get_index_of_marker(marker, table):
        markers_ = [item[0] for item in table]
        return markers_.index(marker) if marker in markers_ else None

    # macht aus dem json-Objekt ein Tabellen-Objekt
    def decode_alignment(map, marker):
        if "mkrFollowingThis" in markers[marker]:
            new_marker = markers[marker]["mkrFollowingThis"]

            # die Map enthält alle annotierbaren Marker. Wenn nach dem letzten Marker wieder "ref" als following eingegeben ist, um die Datei leicht erweiterbar zu machen, dann führt das zu einem unendlichen Loop, der hier durchbrochen wird.
            if not new_marker in map:
                return [[marker, bytes(map[marker], encoding="UTF-8")]] if map[marker] else None

        else:
            return [[marker, bytes(map[marker], encoding="UTF-8")]] if map[marker] else None

        next_line = decode_alignment(map, new_marker)

        if next_line:
            this_line = [marker, bytes(map[marker], encoding="UTF-8")]

            next_line.append(this_line)

            return next_line
        else:
            return [[marker, bytes(map[marker], encoding="UTF-8")]] if map[marker] else None

    # print

    def decode_words(marker, table, prefix, min=0, yy=-1):  # check von min bis xx
        current_index = get_index_of_marker(marker, table)
        if current_index is None:
            return []

        current_row = table[current_index][1]

        xx_ = min
        if yy != -1:
            max = yy
            # print(marker, max, current_row[max:max+1])
            if max < len(current_row):
                if current_row[max] == 32 and "jumps" in markers[marker]:
                    # print("#", min, max, current_row[min:max+1], current_row)
                    return None
        else:
            max = len(current_row)

        if 0 < min < len(current_row) and current_row[min - 1] != 32:
            # wenn das Zeichen (in einem rekursiven Aufruf der Funktion mit bestimmtem min und max) vor dem min kein Leerzeichen ist, können die Zeilen nicht aliniert sein. Deswegen gehen wir vorsichtshalber weiter (→ if following_words == None)
            return None

        spalten = []
        if "jumps" in markers[marker]:

            found = False
            for xx in range(min, max):
                if xx >= len(current_row):
                    # input()
                    # print(marker, min, max, xx, current_row[xx], current_row[xx+1])
                    break

                cond2 = xx == len(current_row) - 1
                # das Ende der Zeile bedeutet auch das Ende eines Wortes, aber die Zeilen sind nicht notwendigerweise gleich lang, weswegen ein anderer Funktionsaufruf benötigt wird
                cond1 = current_row[xx] == 32 and current_row[xx + 1] != 32 if not cond2 else False
                # das Ende von Leerzeichen bedeutet das Ende eines Wortes

                if cond1 or cond2:
                    found = True

                    # hier wird der Eintrag aus der Annotation in die Datenbank eingetragen, strip(" ") um den Zeilenumbruch am Ende zu behalten. Dasselbe gilt aber auch für die Wörterbücher, wo folglich bei jedem Eintrag \n-chars verbleiben
                    current_word = {marker: current_row[xx_:xx + 1].decode("UTF-8").strip()} if cond1 else {
                        marker: current_row[xx_:].decode("UTF-8").strip(" ")}

                    # print(cond1, cond2)
                    # print(current_word)

                    zeilen = []
                    for jump in markers[marker]["jumps"]:
                        next_marker = jump["mkrTo"]

                        following_words = decode_words(next_marker, table, prefix, xx_,
                                                       xx + 1) if cond1 else decode_words(next_marker, table, prefix,
                                                                                          xx_)

                        if following_words is None:
                            if cond2:
                                continue
                            else:
                                break
                        # Es wird nur None zurückgegeben, wenn bei der Alinierung ein Leerzeichen nicht über einem anderen Leerzeichen steht. Das deutet auf eine Spannenannotation in der folgenden Zeile, daher wird die for-Schleife gebrochen und die Zählung xx_ = xx+1 nicht ausgelöst, es wird also bis zum nächsten Leerzeichen gesucht.
                        # Wenn wir uns hingegen am Ende der Zeile befinden, dann muss das current_word noch hinzugefügt werden, ansonsten geht uns der Rest der Zeile verloren, deswegen continue

                        if not following_words:
                            # log.append({**prefix, **current_word})
                            # hier sollte an allen Stellen gemeckert werden, wo die Annotation entweder fehlerhaft ist oder fehlt
                            continue
                        # andere Möglichkeit: beim Überprüfen der Konsistenz loggen

                        following_words = [[following_words[i][y] for i in range(len(following_words))] for y in
                                           range(len(following_words[0]))]

                        for word in following_words:
                            zeilen.append(word)

                    else:  # folgendes wird nicht aufgeführt, wenn zuvor None zurückgegeben wurde
                        # das Array muss symmetrisch sein, damit es rotiert werden kann, das wird durch Verdopplung erreicht. Verdopplung heißt aber, dass eine Spannenannotation vorliegt, wodurch '@' vorgeschaltet werden muss

                        for zz in reversed(range(len(zeilen))):
                            if len(zeilen[zz - 1]) < len(zeilen[zz]):
                                to_append = deepcopy(zeilen[zz - 1])[-1]

                                for key in zeilen[zz - 1][-1]:
                                    zeilen[zz - 1][-1][key] = '@' + zeilen[zz - 1][-1][key]
                                    break  # idR muss nur der erste Eintrag markiert werden

                                zeilen[zz - 1].append(to_append)
                            # Python verweist auf Elemente stets mit Referenzen. Wenn hier keine Kopie eingesetzt wird, werden im Folgenden beim Merging der dictionaries alle Einträge gleich verändert

                        if not zeilen:
                            if cond2:  # wenn am Ende der Zeile immer noch keine Annotationen gefunden wurden, ist die gnaze Zeile leer und wird ohne Annotation in die Datenbank aufgenommen
                                spalten.append([current_word])

                            continue

                        zeilen = [[zeilen[i][y] for i in range(len(zeilen))] for y in range(len(zeilen[0]))]

                        for zz in range(len(zeilen)):
                            zeile = zeilen[zz]

                            if len(zeilen) > 1 and zz < len(zeilen) - 1:
                                zeile.append({key: "@" + current_word[key] for key in current_word})
                            # mit einem @ am Anfang werden Spannenannotationen markiert, damit sie später wieder zusammengesetzt werden können, das letzte wird dabei ausgelassen um das Ende zu markieren (wichtig bei Zeilenumbrüchen)
                            else:
                                zeile.append(current_word)

                            for i in range(len(zeile)):
                                zeile[0].update(zeile[i])

                            zeile = [zeile[0]]  # für logging reasons (???)
                            spalten.append([zeile[0]])

                        xx_ = xx + 1
            if not found:
                if min < yy:
                    return None  # wenn er von vorne bis hinten durchgegangen ist, ohne das cond1 oder cond2 gegriffen haben, muss ein Fehler in der Alignierung vorliegen. Deswegen geht es weiter (→ if following_words == None)
                elif min > yy:
                    return []  # wenn er gar nicht erst durch die Schleife durchgehen konnte, dann weil die Zeile zu kurz ist. Das passiert, wenn das letzte Wort der Mutter-Zeile gar keine Annotation hat.
        else:
            try:
                current_word = {marker: current_row[min:max].decode("UTF-8").strip()} if yy != -1 else {
                    marker: current_row[min:].decode("UTF-8").strip()}
                spalten.append([current_word])
            except UnicodeDecodeError:
                print("Undecodable bytes at", prefix)
                current_word = {marker: current_row[min:].decode("UTF-8").strip()}

        for spalte in spalten:
            for word in spalte:
                word.update(prefix)

        # input(spalten)

        return spalten  # if spalten != [] else [] # ;)

    new_marker = markers[marker]["mkrFollowingThis"]

    if "jumps" not in markers[marker]:
        # vgl. oben. mit dem Vorhandensein wird zwischen Id- und Record-Markern und zu interlinearisierenden Einträgen unterschieden. Bei den Text-Datenbanken ist das nativ von Toolbox vorgesehen
        for element in map[marker]:  # dictionarys für Knoten
            if element is None: continue
            prefix.update({marker: element})
            decode_toolbox_json(map[marker][element], new_marker, prefix)
        return

    ref_marker = ""
    for element in map:  # listen für ref-Gruppen
        if Is.do_filter and "ref" in prefix.keys():
            # print(prefix["ref"] + " " + str(is_in_subpart(prefix["ref"])))
            sub_part = is_in_subpart(prefix["ref"], Is.do_filter)
            if not sub_part:
                continue

        table = decode_alignment(element, marker)
        if table is None:
            continue

        decoded_table = [ddict for llist in decode_words(marker, table, prefix) for ddict in llist]

        if Is.do_reload:
            if len(set([ddict["ref"] for ddict in decoded_table])) > 1:
                input(decoded_table)

            ii = 0
            if ref_marker and ref_marker != decoded_table[0]["ref"] or ref_marker == "":
                uu = 0

            ref_marker = decoded_table[0]["ref"]
            if ref_marker in raw_xml:
                xml_tx = raw_xml[ref_marker]
                xml_lst = xml_tx.split(" ")
            else:
                break

            # print(ref_marker)
            # print(" ".join([ddict["tx"] for ddict in decoded_table]).strip())
            # print(xml_tx)

            while ii < len(decoded_table):
                toolbox_tx = decoded_table[ii]["tx"]
                toolbox_tx = re.sub(" +", " ", toolbox_tx)

                tx2 = xml_lst[uu]
                while tx2 in ["„", "‚", "“", "‘"]:
                    uu += 1
                    tx2 = xml_lst[uu]

                tx1 = re.sub(' ', ' ', toolbox_tx)
                if tx1[-1] == "\n":
                    tx2 += "\n"

                tx2_joints = tx2.count("⸗") + tx2.count("-") + tx2.count("=")
                tx1_breaks = tx1.count(" ")

                # print(ii, tx1, tx1 == tx2, tx2)
                if not tx1 == tx2:
                    if ii + 1 == len(decoded_table):  #
                        tx2 = " ".join(xml_lst[uu:])
                        if tx1[-1] == "\n":
                            tx2 += "\n"

                        if decoded_table[ii]["tx"] == tx2:
                            # es fehlen lediglich Annotationen, der Rest ist identisch
                            uu += tx2.count(" ")
                            break
                        elif uu + tx1_breaks + 1 < len(xml_lst) and decoded_table[ii]["tx"] == xml_lst[
                            uu + tx1_breaks + 1] + "\n":
                            # und nicht auf der nächsten Zeile noch ein anderes Wort dazwischen gerutscht ist (passiert bei Zeilen mit Zwei wörtern, wenn darüber ein Zeilenumbruch eingefügt wird),
                            print(ref_marker, "initial surplus word. Possible line break?")
                            uu += 1
                            continue

                    # das Wort oder die Phrase ist äquivalent zu allem, was folgt. Dadurch ist ein index-overflow im Folgenden ausgeschlossen

                    elif tx1[0] == "@":  # Spannenannotationen werden übersprüngen und ggf. rückwirkend korrigiert
                        ii += 1
                        continue

                    elif tx1 == " ".join(xml_lst[
                                         uu:uu + tx1_breaks + 1]):  # in diesem Fall stimmen beide Teile überein, sobald man die Leerzeichen berücksichtigt. Das ist der Fall, wenn Nobreak-Spaces oder Leerzeichen im Text vorhanden sind. Letzeteres kann bei unvollständigen Annotationen auftreten
                        tx2 = " ".join(xml_lst[uu:uu + tx1_breaks + 1])
                        if tx1[-1] == "\n":
                            tx2 += "\n"

                        uu += tx1_breaks + 1
                        ii += 1
                        continue

                    elif decoded_table[ii + 1]["tx"] == xml_lst[
                        uu + 1]:  # wenn das nächste Wort übereinstimmt, liegt hier Äquivalenz vor und es kann ausgetauscht werden
                        tx2 = xml_lst[uu]

                    elif decoded_table[ii + 2]["tx"] == xml_lst[
                        uu + tx1_breaks + 1]:  # das übernächste Wort stimmt überein
                        if not tx1_breaks:  # wenn keine Leerzeichen oder Nbksp vorhanden sind, die auf eine Zusammenrückung in der Korrektur hinweisen, wurde wahrscheinlich ein Wort gelöscht
                            print("deleted '" + decoded_table[ii]["tx"] + "' at index", ii, "in", ref_marker)
                            decoded_table = decoded_table[:ii] + decoded_table[ii + 1:]

                            continue

                        else:
                            print(ref_marker)
                            print(" ".join([ddict["tx"] for ddict in decoded_table]))
                            print(xml_tx)
                            print(ii, uu, tx1_breaks)
                            print("breaks")
                            input()

                    elif decoded_table[ii]["tx"] == xml_lst[
                        uu + tx1_breaks + 1]:  # Das Wort wurde in der XML-Datei gelöscht
                        if not tx1_breaks:  # wenn keine Leerzeichen oder Nbksp vorhanden sind, die auf eine Zusammenrückung in der Korrektur hinweisen, wurde wahrscheinlich ein Wort gelöscht
                            if ii == 0:  # am Anfang der Zeile deuten Löschungen auf Zeilenumbrüche hin
                                print(ref_marker, "initial surplus word. Possible line break?")

                            else:
                                print(ref_marker)
                                print(" ".join([ddict["tx"] for ddict in decoded_table]))
                                print(xml_tx)
                                print(ii, uu, tx1_breaks)
                                print("deleted in-place")

                            input()

                            uu += 1
                            continue

                        else:
                            print(ref_marker)
                            print(" ".join([ddict["tx"] for ddict in decoded_table]))
                            print(xml_tx)
                            print(ii, uu, tx1_breaks)
                            print("breaks in-place")
                            input()


                    else:
                        print(ref_marker)
                        print(" ".join([ddict["tx"] for ddict in decoded_table]).strip())
                        print(xml_tx)

                        print("Please check the source file and try again")

                        input()
                        break

                    if ("|" not in tx2) and ("|" in tx1):
                        print(ref_marker, "please add a linebreak to the xml")
                        return

                    existing_annotations = [entry for entry in db_words[markers[marker]["jumps"][0]["dbtyp"]] if
                                            entry[marker].strip("\n").replace(' ', ' ') == toolbox_tx.strip("\n")]
                    if existing_annotations:
                        if not [entry for entry in db_words[markers[marker]["jumps"][0]["dbtyp"]] if
                                entry[marker].strip("\n").replace(' ', ' ') == tx1.strip("\n")]:
                            for ann in existing_annotations:
                                new_ann = dict(copy(ann), **{marker: tx1.replace(' ', ' ').strip("\n") + "\n"})
                                db_words[markers[marker]["jumps"][0]["dbtyp"]].append(new_ann)
                                print("new", new_ann)

                    yy = 1  # Behandlung von Spannen
                    while yy < ii and "@" + decoded_table[ii]["tx"] == decoded_table[ii - yy]["tx"]:
                        decoded_table[ii - yy]["tx"] = "@" + tx2

                        print("@@@")
                        input(" ".join([ddict["tx"] for ddict in decoded_table]))

                    if not ("|" in tx2 or len(decoded_table) == 1):
                        print(ref_marker, decoded_table[ii]["tx"], "→", tx2)
                    decoded_table[ii]["tx"] = tx2

                uu += 1
                ii += 1

        # input()

        for dictt in decoded_table:
            # wenn die Wörter hier korrigiert werden, wird die Laufzeit um mehrere Stunden verkürzt
            words.extend(check_word_for_consistency(dictt, marker))


def check_word_for_consistency(word, marker):
    def check_word_for_consistency_(word, marker):
        global spannenindex
        if not marker in spannenindex:
            spannenindex.update({marker: 0})

        def strip_plus(string):
            if string is not None:
                return string.strip('@').strip().lower()  # doppeltes Strip wegen Spannenannotation und \n-Markern
            else:
                return None

        # diese Funktion läuft durch alle Marker, die Teil einer jump-Funktion sind, also die erste Zeile in einer Datenbank sind
        # print(marker + " " + str(word[marker]) + " " + str(word))
        for jump in markers[marker]["jumps"]:
            jumpFrom = jump["mkr"]
            jumpTo = jump["mkrTo"]
            jumpToDb = jump["dbtyp"]
            jumpOut = jump["mkrOut"]

            word[marker] = re.sub("( | )+", " ", word[
                marker])  # für den Import in ANNIS müssen Zeilenzusammenrückungen durch Nobreak-Spaces gelöst werden. Da nur jump-Marker in dieser Funktion durchgenommen werden, wird die Integrität der Toolboxdateien nicht gefährdet.

            # strip(), weil in der Datenbank überall und in der Annotationsdatei an Zeilenumbrüchen noch \n-chars sind
            db_words_ = [db_word for db_word in db_words[jumpToDb] if
                         strip_plus(db_word[jumpFrom]) == strip_plus(word[marker])]

            database_annotations = [db_word[jumpOut] for db_word in db_words_ if db_word[jumpOut] is not None]
            database_annotations = [dbw.strip() for dba in database_annotations for dbw in
                                    dba.split(jump['GlossSeparator'])]
            if "jumps" in markers[jumpTo]:
                database_annotations = [dba.split() for dba in database_annotations]
            # für alle Datenbankeinträge: in allen durch den GlossSeparator abgetrennten Glossen: alle durch Leerzeichen getrennten Teilwörtern sind eligible für die Interlinearisierung auf der nächsten Zeile
            else:
                database_annotations = [[dba] for dba in database_annotations]
            # bei Zeilen, die nicht weiter segmentiert werden, sind die Leerzeichen Teil des Annotationswertes und dürfen nicht gesplittet werden.

            database_annotations.sort()
            database_annotations = list(el for el, _ in itertools.groupby(database_annotations))
            database_annotations_ = [[strip_plus(dbw)] for dba in database_annotations for dbw in dba]

            # falls mal wieder Debugging nötig wird
            # print(database_annotations)
            # print(jumpTo, jumpTo in word, any([strip_plus(word[jumpTo]) in dba for dba in database_annotations_ if jumpTo in word]))
            # input()

            # wenn bei zu interlinearisierenden Teilwörtern auf der folgenden Ebene keine Suffix oder Lisaisons-Marker vorhanden sind, wird beim Pepper-Export nach ANNIS die Alignierung falsch abgebildet werden. Das brauch ich für den Excel-Iporter nicht.
            # if any([isinstance(db, list) and len(db) > 1 for db in database_annotations_]):
            #	if not any([(string[-1] in ['=', '-'] or string[0] in ['_', '=', '-']) for dba in database_annotations_ for string in dba]) and word[marker][0] != '@':
            #		log.append({**{"tofix" : "alignment (2): " + jumpTo}, **word})
            # else:
            #	#print(database_annotations_)
            #	if any([(string[-1] in ['=', '-'] or string[0] in ['=', '-']) for dba in database_annotations_ for string in dba]):
            #		log.append({**{"tofix" : "alignment (1): " + jumpTo}, **word})

            # Die Annotation ist eindeutig
            if len(database_annotations) == 1:
                if spannenindex[marker] >= len(database_annotations[0]):
                    from_database = database_annotations[0][-1]
                    log.append({**{"tofix": "possible duplicate"}, **word})
                    print("possible duplicate: ", database_annotations[0], marker, spannenindex[marker], str(word))

                else:
                    from_database = database_annotations[0][spannenindex[marker]]

                # Die Annotation fehlt und wird automatisch hinzugefügt
                if not jumpTo in word:
                    word.update({jumpTo: from_database})
                    log.append({**{"fixed": jumpTo}, **word})
                # Es gibt eine Annotation, die allerdings nicht mit der Datenbank übereinstimmt
                elif not any(strip_plus(word[jumpTo]) in dba for dba in database_annotations_):
                    word[jumpTo] = from_database
                    log.append({**{"fixed": jumpTo}, **word})
                else:
                    # wenn alles stimmt, müssen nur Zeilenumbrüche vermieden werden
                    if word[jumpTo][-1] == "\n":
                        word[jumpTo] = word[jumpTo][:-1]

            # Die Annotation ist nicht eindeutig
            else:
                # Die Annotation fehlt
                if not jumpTo in word:
                    log.append({**{"tofix": jumpTo + ": " + str(database_annotations)}, **word})
                    return word
                # Die Annotation ist vorhanden, nichts muss getan werden
                elif [strip_plus(word[jumpTo])] in database_annotations_:
                    pass

                # Die hinterlegte Annotation ist kein Substring eines Eintrags in der Datenbank
                elif not any([strip_plus(word[jumpTo]) in dbw for dba in database_annotations_ for dbw in dba]):
                    log.append({**{"tofix": jumpTo}, **word})

                # Die hinterlegte Annotation ist Substring eines Eintrags in der Datenbank
                else:
                    database_annotations_ = [[dbw] for dba in database_annotations for dbw in dba if
                                             strip_plus(word[jumpTo]) in dbw]

                    if len(database_annotations_) == 1:
                        # das sollte eigentlich nicht vorkommen
                        if spannenindex[marker] >= len(database_annotations_[0]):
                            from_database = database_annotations_[0][-1]
                            log.append({**{"tofix": "possible duplicate"}, **word})
                            print("possible duplicate in multiple: ", database_annotations_[0], marker,
                                  spannenindex[marker], str(word))

                        else:
                            from_database = database_annotations_[0][spannenindex[marker]]

                        word[jumpTo] = from_database
                        log.append({**{"fixed": jumpTo}, **word})

            if len(word[marker]):
                if word[marker][0] == '@':
                    spannenindex[marker] += 1
                elif word[marker][0] != '@' and spannenindex[marker] != 0:
                    # print(spannenindex, word)
                    for marker in spannenindex:
                        spannenindex[marker] = 0
            else:
                print("what?", word)

            if "jumps" in markers[jumpTo]:
                word = check_word_for_consistency_(word, jumpTo)

        return word

    def automatically_annotate(word, marker):
        list = []
        for splitt in word[marker].split(" "):
            new_word = word.copy()
            new_word[marker] = splitt
            list.append(check_word_for_consistency_(new_word, marker))

        return list

    if Is.do_check and not (Is.ignore_numbers and re.match("^\d+\.?$", word[marker].strip())):
        if len(word) > 4:  # fName, id, ref, tx → keine Annotationen
            return [check_word_for_consistency_(word, marker)]
        else:
            return automatically_annotate(word, marker)
    else:
        return [word]


# speichert die geladenen und bearbeiteten Daten in einem lokalen Unterordner im Toolbox-Format ab
def list_to_toolbox(words, root_marker):
    def repl(m):
        return " " * len(m.group())

    def last_char_after_strip(string, char):
        string = string.strip(" ")
        if len(string) > 0:
            if string[-1] == char:
                return True
        return False

    current_block = {}
    in_spann = {}

    def compose_block(word, marker):
        new_block = {}

        while True:
            if "mkrFollowingThis" in markers[marker] and markers[marker]["mkrFollowingThis"] in new_block:
                # in manchen Dateien ist als following auf den Übersetzungsmarker wieder ref eingegeben, um die Datei leicht zu erweitern. Das führt zu einem unendlichen Loop, der hier unterbrochen wird. Weil der "ref" Marker selbst nicht in den hier verarbeiteten Daten vorhanden ist, muss einen weiteren Schritt vorausgeschaut werden, damit der tx-Marker entdeckt wird, ansonsten wird "ref" als Annotation reproduziert.
                break

            new_block[marker] = bytes(word[marker], encoding="UTF-8") if marker in word else b''

            if "mkrFollowingThis" in markers[marker]:
                marker = markers[marker]["mkrFollowingThis"]

            else:
                break

        target_length = len(max(new_block.values(), key=len)) + 1  # wir brauchen ein Leerzeichen als Trenner

        for key in new_block:
            if len(new_block[key]) > 0 and new_block[key][0:1] != b'@' or len(new_block[key]) == 0:

                if not key in in_spann:
                    in_spann[key] = False

                new_value = new_block[key]
                new_value += b' ' * (target_length - len(new_value))
                new_value = new_value.decode("UTF-8")

            elif len(new_block[key]) > 0 and new_block[key][0:1] == b'@':
                new_value = b'@' * target_length
                new_value = new_value.decode("UTF-8")

                if not key in in_spann:
                    in_spann[key] = True

            if not key in current_block:
                current_block[key] = new_value
            else:
                current_block[key] += new_value

    def dict_to_text(dict):
        string = ""
        for key in dict:
            if not dict[key] == "":
                string += "\\" + key + " " + dict[key].strip() + "\n"
                dict[key] = ""

        if string == "":
            return ""

        # string += "\n"

        # Spannenannotation werden aufgelöst
        string = re.sub("(@+)([^ \@\n]+)", "\g<2>\g<1>", string)
        string = re.sub("@+", repl, string)

        return string

    def current_file_path():
        path = os.path.join(os.path.basename(Paths.toolbox_folder), current_file_name + ".txt")
        # path = os.path.join(toolbox_folder_path, current_file_name + "_ed.txt")
        return path

    current_file_name = ""
    current_file_content = "\\_sh v3.0  621  Text\n"

    print("\nSchreibe Dateien:")

    current_record = {}
    for entry in words:
        if not Is.in_one_file:
            if current_file_name == "":
                current_file_name = entry["fName"]
            elif entry["fName"] != current_file_name and not current_file_name == "":
                open(current_file_path(), "w", encoding="utf-8").write(current_file_content)
                print(current_file_name + " written")

                current_file_name = entry["fName"]
                current_file_content = "\\_sh v3.0  621  Text\n"
        else:
            if current_file_name == "":
                current_file_name = os.path.basename(Paths.toolbox_folder)

        marker = root_marker
        while "mkrFollowingThis" in markers[
            marker]:  # iteriert durch die Spalten id und ref, solange bis die Annotation erreicht ist
            if not "jumps" in markers[marker]:  # id und ref
                if not marker in current_record:
                    current_record.update({marker: entry[marker]})
                    current_file_content += "\n\\" + marker + " " + entry[marker] + "\n"
                else:
                    if current_record[marker] != entry[marker]:
                        current_file_content += dict_to_text(current_block)

                        current_file_content += "\n\\" + marker + " " + entry[marker] + "\n"
                        current_record[marker] = entry[marker]

                marker = markers[marker]["mkrFollowingThis"]
            else:  # Annotation ist erreicht, neues Protokoll
                compose_block(entry, marker)

                if last_char_after_strip(current_block[marker], "\n"):
                    current_file_content += dict_to_text(current_block)

                break  # die restlichen Spalten wurden ja schon in der compose_block Funktion durchiteriert

    ###
    # scheint obsolet zu sein
    ###
    current_file_content += dict_to_text(current_block)

    open(current_file_path(), "w", encoding="utf-8").write(current_file_content)
    print(current_file_name + " written")


def read_original(read_path):
    def open_xml(path):
        with open(path, "r", encoding="utf-8") as f:
            return re.sub("[\r\n]+?( +)?", "", f.read())

    def intify(str):
        return int(re.search("\d+", str).group(0))

    def return_line_nr(nr):
        if intify(nr) < 10:
            return "0" + str(nr)
        else:
            return str(nr)

    def return_page_nr(nr):
        if intify(nr) < 10:
            return "00" + str(nr)
        elif intify(nr) < 100:
            return "0" + str(nr)
        else:
            return str(nr)

    def split_next_line(soup):
        splitted = soup.text.split(" ", 1)
        if len(splitted) > 1:
            return splitted[1]
        else:
            return ""

    raw = open_xml(read_path)
    soup = BeautifulSoup(raw, "html.parser")

    doc_title = soup.document["title"]

    reference, text = "", ""
    for page in soup.document.find_all("lpp"):
        for line in page.find_all("z"):
            if line["nr"] in ["re", "title"]:
                continue

            if reference and text:
                if text[-1] in ["⸗", "-", "_"]:
                    # text = text.strip("_")
                    text += "|" + line.text.split(" ", 1)[0]
                    line.string.replace_with(split_next_line(line))

                raw_xml[reference] = text

            text = line.text
            reference = "{title}_{page}.{line}".format(title=doc_title, page=return_page_nr(page["nr"]),
                                                       line=return_line_nr(line["nr"]))

if __name__ == '__main__':
    init_globals_from_args(sys.argv[1:])

    # load Toolbox-Data

    # wird verwendet, um die gelesenen Marker zu filtern.
    # do_filter wird als Liste von allen Markern definiert, nach denen gefilert wird.

    if Is.do_filter:
        filters = getFilterFrom(Paths.filter)

    # Generation der Listen mit den Dateien
    print("Lade Datenbanken...")

    toolbox_folder = list(os.walk(Paths.toolbox_folder))

    other_files = []
    typ_files = {}
    for path, _, files in toolbox_folder:
        for file in files:
            if is_valid_file(file):
                other_files.append(os.path.join(path, file))

            if file[-3:] == "typ":
                typ_files[file[:-4]] = os.path.join(path, file)

    types = {}
    for typ in typ_files:
        file_text = open(typ_files[typ], "r", encoding="UTF-8").read().replace("\\", "\\\\")
        type_name = re.search("\\\\\+DatabaseType (\S+)", file_text).group(1)

        mkr_record = re.search("\\\\mkrRecord (\S+)", file_text).group(1)
        gloss_seperator = re.search("\\\\GlossSeparator (\S)", file_text).group(1) if re.search("GlossSeparator",
                                                                                                file_text) else None

        markers = {}
        mkr_texts = re.findall("\\\\\+mkr [\s\S]+?\\\\-mkr", file_text)
        for mkr_text in mkr_texts:
            keys = {}
            for match in re.findall("(?<=[\\\\+])([a-zA-Z]+) (.+?)\\n", mkr_text):
                keys[match[0]] = match[1]
                if match[0] == "mkr":
                    key = match[1]

            markers[key] = keys

        jmp_texts = re.findall("\\\\\+intprc [\s\S]+?\\\\-intprc", file_text)

        if not jmp_texts:
            print("No jumps in marker {}. You might add lookup function?".format(key))

        for jmp_text in jmp_texts:

            keys = {}
            for match in re.findall("(?<=[\\\\+])([a-zA-Z]+) (.+?)\\n", jmp_text):
                keys[match[0]] = match[1]
                if match[0] == "mkr":
                    key = match[1]

            if "jumps" in markers[key]:
                markers[key]["jumps"].append(keys)
            else:
                markers[key]["jumps"] = [keys]

        types[type_name] = [{"mkrRecord": mkr_record, "GlossSeparator": gloss_seperator, "markers": [markers]}]

    # mit der Liste der Sprachtypen soll das Verhalten von Toolbox nachgeahmt werden, nur vordefinierte Zeichen zu zählen
    # lng_files = [os.path.join(path, file) for path, dirs, files in toolbox_folder for file in files if file[-3:] == "lng"]
    # languages = {}
    # for lng in lng_files:
    # file_text = open(lng, "r", encoding="UTF-8").read().replace("\\", "\\\\")
    # input(file_text)

    # das sind dann die Dateien, aus denen tatsächlich Informationen extrahiert werden
    databases = [[file, open(file, "r", encoding="UTF-8").readlines()[0].split(" ", 4)[4].strip()] for file in
                 other_files]
    # Bsp.: "\_sh v3.0  621  Text". Wichtig: doppelte Leerzeichen
    db_files = [tuple for tuple in databases if tuple[1] != "Text"]
    text_files = [tuple for tuple in databases if tuple[1] == "Text"]

    # hier werden die Wörterbücher geladen, damit in Realtime verglichen werden kann, was konsistent ist und was nicht

    for db_file in db_files:
        file_path, typ = db_file[0], db_file[1]

        with open(file_path, "r", encoding="UTF-8") as file:
            file_text = file.read()

        if not file_text[-1] == "\n":
            file_text += "\n"

        if typ not in types:
            continue

        root_marker = types[typ][0]["mkrRecord"]
        markers = types[typ][0]["markers"][0]

        map = decode_toolbox_map(file_text, root_marker)
        df = pandas.DataFrame.from_records(map)
        dpl = df[df.duplicated(keep='first')]
        if not dpl.empty:
            print(typ, "has duplicates:")
            print(dpl.to_string())
            input("press enter to continue")

        db_words[typ] = map

    # bring the action
    print("\nLese Dateien:")
    for text_file in text_files:
        file_path, typ = text_file[0], text_file[1]

        with open(file_path, "r", encoding="UTF-8") as file:
            print(file_path)
            file_text = file.read()

        if not file_text[-1] == "\n":
            file_text += "\n"

        root_marker = types[typ][0]["mkrRecord"]
        markers = types[typ][0]["markers"][0]

        map = decode_toolbox_map(file_text, root_marker)

        filename = os.path.basename(file_path)
        filename = filename[:-4] if ".txt" in filename else filename
        decode_toolbox_json(map, root_marker, {"fName": filename})

    if Is.reexport:
        list_to_toolbox(words, root_marker)

    df = pandas.DataFrame.from_records(words)
    df = df.replace(r'\n', '', regex=True)
    df.to_csv(Paths.output_folder, sep=';', encoding="UTF-8-SIG", index=False, header=True)

    if Paths.excel_export:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "AnnotationData"

        for row in dataframe_to_rows(df, index=False, header=True):
            ws.append(row)

        for col in ws.iter_cols():
            merge_first = False
            for cell in col:
                cc = cell.column

                if col[0].value in ["fName", "id", "ref"]:
                    if cell.row > 1 and cell.value and str(ws.cell(column=cc, row=cell.row - 1).value) != cell.value:
                        if not merge_first:
                            merge_first = cell.row
                            merge_last = False
                        else:
                            merge_last = cell.row - 1

                            ws.merge_cells(start_column=cc, end_column=cc, start_row=merge_first, end_row=merge_last)

                            merge_first = cell.row

                else:
                    if not merge_first and cell.value and str(cell.value)[0] == "@":
                        merge_first = cell.row
                    # cell.value = None
                    elif merge_first and cell.value[0] != "@":
                        merge_last = cell.row
                        ws.cell(row=merge_first, column=cc).value = cell.value

                        ws.merge_cells(start_column=cc, end_column=cc, start_row=merge_first, end_row=merge_last)

                        merge_first = False
                        merge_last = False

            if merge_first:
                merge_last = col[-1].row
                ws.merge_cells(start_column=cc, end_column=cc, start_row=merge_first, end_row=merge_last)

        meta_file_path = os.path.join(Paths.toolbox_folder,
                                      os.path.splitext(os.path.basename(Paths.excel_export))[0] + ".json")
        if os.path.isfile(meta_file_path):
            ms = wb.create_sheet("MetaData")

            meta_text = open(meta_file_path, 'r').read()
            meta_info = json.loads(meta_text, ensure_ascii=False)

            for key, value in meta_info.items():
                ms.append([key, value])

        while True:
            try:
                wb.save(Paths.excel_export)
                break
            except PermissionError:
                input("please close " + Paths.excel_export)

        print("\n" + Paths.excel_export, "written")

    if log:
        df = pandas.DataFrame.from_records(log)
        df = df.replace(r'\n', '', regex=True)

        print("\nlength of log:")
        print(df["fName"].value_counts())
        df.to_csv(Paths.log_file, sep=';', encoding="UTF-8-SIG", index=False, header=True)
    else:
        print("no log")

    print(len(words), "words")
